#!/usr/bin/env python3
"""Converte um backup XLSX do Google Sheets em snapshot JSON sem alterar a origem."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def serialize(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def extract_sheet(worksheet) -> list[dict[str, Any]]:
    row_iterator = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iterator)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        return []
    if any(not header for header in headers):
        raise ValueError(f"A aba {worksheet.title!r} possui cabeçalho vazio entre colunas preenchidas.")
    if len(set(headers)) != len(headers):
        raise ValueError(f"A aba {worksheet.title!r} possui cabeçalhos duplicados.")

    rows: list[dict[str, Any]] = []
    for source_row, values in enumerate(row_iterator, start=2):
        relevant = list(values[: len(headers)])
        if not any(value not in (None, "") for value in relevant):
            continue
        item = {header: serialize(value) for header, value in zip(headers, relevant)}
        item["_sourceRow"] = source_row
        rows.append(item)
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Extrai um backup XLSX para o formato seguro de migração.")
    parser.add_argument("input", type=Path, help="Arquivo .xlsx exportado da planilha de backup")
    parser.add_argument("output", type=Path, nargs="?", default=Path("scripts/migration/source/backup.json"))
    parser.add_argument("--spreadsheet-id", default="", help="ID da planilha de origem para auditoria")
    args = parser.parse_args()

    if not args.input.is_file():
        raise FileNotFoundError(f"Arquivo não encontrado: {args.input}")

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    try:
        snapshot = {
            "sourceSpreadsheetId": args.spreadsheet_id,
            "exportedAt": datetime.now().astimezone().isoformat(),
            "sheets": {worksheet.title: extract_sheet(worksheet) for worksheet in workbook.worksheets},
        }
    finally:
        workbook.close()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    counts = {name: len(rows) for name, rows in snapshot["sheets"].items()}
    print(json.dumps({"success": True, "output": str(args.output), "counts": counts}, ensure_ascii=False))


if __name__ == "__main__":
    main()
